"""
CSS Amazon Monthly P&L — Streamlit Web App
Upload the 3 Amazon source files and download the finished Excel workbook.
"""

import io
import csv
import streamlit as st
from collections import defaultdict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# ── constants ────────────────────────────────────────────────────────────────
MONTH_NAMES = ['', 'January', 'February', 'March', 'April', 'May', 'June',
               'July', 'August', 'September', 'October', 'November', 'December']
MONTH_LETTER = {1:'Ja',2:'F',3:'M',4:'A',5:'My',6:'Jn',7:'Jl',8:'Au',9:'S',10:'O',11:'N',12:'D'}

HEADER_FILL  = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
TITLE_FONT   = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True)
TOTAL_FONT   = Font(bold=True)
NUMBER_FMT   = '#,##0.00;[Red]-#,##0.00'
INT_FMT      = '#,##0'

# ── helpers ───────────────────────────────────────────────────────────────────
def to_float(s):
    if s is None or s == '': return 0.0
    if isinstance(s, (int, float)): return float(s)
    return float(str(s).strip().strip('"').replace(',', '') or 0)

def to_int(s):
    if s is None or s == '': return 0
    return int(float(str(s).strip().strip('"') or 0))

def parse_amazon_dt(s):
    if not s: return None
    s = s.strip().strip('"')
    try:
        return datetime.strptime(s, '%d %b %Y %H:%M:%S UTC')
    except ValueError:
        return None

def read_text(uploaded_file):
    """Return file bytes decoded as UTF-8 (with BOM fallback)."""
    raw = uploaded_file.read()
    for enc in ('utf-8-sig', 'utf-8', 'latin-1'):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode('latin-1')

def load_all_orders(uploaded_file):
    text = read_text(uploaded_file)
    reader = csv.DictReader(io.StringIO(text), delimiter='\t')
    return list(reader)

def load_transactions(uploaded_file):
    text = read_text(uploaded_file)
    lines = text.splitlines(keepends=True)
    body  = ''.join(lines[9:])          # skip 9 description rows
    reader = csv.DictReader(io.StringIO(body))
    return list(reader)

def load_deferred(uploaded_file):
    text = read_text(uploaded_file)
    lines = text.splitlines(keepends=True)
    body  = ''.join(lines[7:])          # skip 7 description rows
    reader = csv.DictReader(io.StringIO(body))
    return list(reader)

# ── core P&L builder (returns workbook + summary dict) ───────────────────────
def build_workbook(year, month, all_orders, transactions, deferred):
    month_name = MONTH_NAMES[month]
    tab_prefix = MONTH_LETTER[month]

    # scope
    target_order_ids = set()
    shipped_order_ids = set()
    cancelled_order_ids = set()
    units_by_status = defaultdict(int)
    total_line_items = 0

    for row in all_orders:
        oid    = row.get('amazon-order-id', '')
        status = row.get('order-status', '')
        qty    = to_int(row.get('quantity', 0))
        target_order_ids.add(oid)
        total_line_items += 1
        units_by_status[status] += qty
        if status == 'Shipped':    shipped_order_ids.add(oid)
        elif status == 'Cancelled': cancelled_order_ids.add(oid)

    # cross-ref transactions
    matched_orders, matched_refunds = [], []
    period_service_fees, period_fba_fees, period_adjustments, period_amazon_fees = [], [], [], []
    target_order_ids_seen_in_tx = set()

    for row in transactions:
        ttype        = row.get('type', '')
        oid          = row.get('order id', '')
        dt           = parse_amazon_dt(row.get('date/time', ''))
        posted_month = dt.month if dt else None
        posted_year  = dt.year  if dt else None
        is_target    = (posted_year == year and posted_month == month)

        if ttype == 'Order' and oid in target_order_ids:
            row['_source'] = 'primary' if is_target else 'secondary'
            matched_orders.append(row)
            target_order_ids_seen_in_tx.add(oid)
        elif ttype == 'Refund' and oid in target_order_ids:
            row['_source'] = 'primary' if is_target else 'secondary'
            matched_refunds.append(row)
        elif ttype == 'Service Fee'       and is_target: period_service_fees.append(row)
        elif ttype == 'FBA Inventory Fee' and is_target: period_fba_fees.append(row)
        elif ttype == 'Adjustment'        and is_target: period_adjustments.append(row)
        elif ttype == 'Amazon Fees'       and is_target: period_amazon_fees.append(row)

    # cross-ref deferred
    matched_deferred, deferred_order_ids = [], set()
    for row in deferred:
        oid = row.get('order id', '')
        if row.get('type') == 'Order' and oid in target_order_ids and oid not in target_order_ids_seen_in_tx:
            matched_deferred.append(row)
            deferred_order_ids.add(oid)

    missing = shipped_order_ids - (target_order_ids_seen_in_tx | deferred_order_ids)

    # aggregations
    def scol(rows, col): return sum(to_float(r.get(col, 0)) for r in rows)

    secondary_orders = [r for r in matched_orders if r['_source'] == 'secondary']
    secondary_units  = sum(to_int(r.get('quantity', 0)) for r in secondary_orders)
    secondary_value  = scol(secondary_orders, 'total')

    order_product_sales = scol(matched_orders, 'product sales')
    order_postage       = scol(matched_orders, 'postage credits')
    order_giftwrap      = scol(matched_orders, 'gift wrap credits')
    order_product_tax   = scol(matched_orders, 'product sales tax')
    order_shipping_tax  = scol(matched_orders, 'shipping credits tax')
    order_promo         = scol(matched_orders, 'promotional rebates')
    order_selling_fees  = scol(matched_orders, 'selling fees')
    order_fba_fees      = scol(matched_orders, 'fba fees')
    order_other_fees    = scol(matched_orders, 'other transaction fees')

    deferred_product = scol(matched_deferred, 'product sales')
    deferred_selling = scol(matched_deferred, 'selling fees')
    deferred_fba     = scol(matched_deferred, 'fba fees')
    deferred_other   = scol(matched_deferred, 'other transaction fees')

    refund_product = scol(matched_refunds, 'product sales')
    refund_postage = scol(matched_refunds, 'postage credits')
    refund_promo   = scol(matched_refunds, 'promotional rebates')
    refund_selling = scol(matched_refunds, 'selling fees')
    refund_fba     = scol(matched_refunds, 'fba fees')
    refund_other   = scol(matched_refunds, 'other transaction fees')

    gross_revenue    = order_product_sales + order_postage + order_giftwrap + order_product_tax + order_shipping_tax
    total_promo      = order_promo
    total_refund_net = refund_product + refund_postage + refund_promo + refund_selling + refund_fba + refund_other
    net_revenue      = gross_revenue + total_promo + total_refund_net
    total_order_fees = order_selling_fees + order_fba_fees + order_other_fees

    def by_desc(rows):
        d = defaultdict(float)
        for r in rows:
            d[r.get('description','').strip().strip('"')] += to_float(r.get('total', 0))
        return d

    service_by_desc    = by_desc(period_service_fees)
    fba_inv_by_desc    = by_desc(period_fba_fees)
    adjust_by_desc     = by_desc(period_adjustments)
    amazon_fees_by_desc = by_desc(period_amazon_fees)

    total_service     = sum(service_by_desc.values())
    total_fba_inv     = sum(fba_inv_by_desc.values())
    total_adjustments = sum(adjust_by_desc.values())
    total_amazon_fees = sum(amazon_fees_by_desc.values())
    net_contribution  = net_revenue + total_order_fees + total_service + total_fba_inv + total_adjustments + total_amazon_fees

    # SKU rollup
    sku_data = defaultdict(lambda: {'units':0,'refund_units':0,'product_sales':0.0,
                                     'promo':0.0,'refund_amt':0.0,'selling_fees':0.0,'fba':0.0,'other':0.0})
    for r in matched_orders:
        sku = r.get('sku','').strip().strip('"') or 'NO-SKU'
        sku_data[sku]['units']         += to_int(r.get('quantity',0))
        sku_data[sku]['product_sales'] += to_float(r.get('product sales',0))
        sku_data[sku]['promo']         += to_float(r.get('promotional rebates',0))
        sku_data[sku]['selling_fees']  += to_float(r.get('selling fees',0))
        sku_data[sku]['fba']           += to_float(r.get('fba fees',0))
        sku_data[sku]['other']         += to_float(r.get('other transaction fees',0))
    for r in matched_refunds:
        sku = r.get('sku','').strip().strip('"') or 'NO-SKU'
        sku_data[sku]['refund_units'] += abs(to_int(r.get('quantity',0)))
        sku_data[sku]['refund_amt']   += to_float(r.get('product sales',0))

    # ── build workbook ──────────────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)

    # Contents
    ws = wb.create_sheet('Contents')
    ws['A1'] = f'CSS Amazon UK - {month_name} {year}'
    ws['A1'].font = TITLE_FONT
    ws['A4'] = 'CONTENTS'; ws['A4'].font = SECTION_FONT
    rows_c = [
        (None,None),(('Contents','This page')),(('Reports Used','Why three reports are needed')),
        (None,None),(f'{month_name.upper()} {year}',None),
        (f'{tab_prefix}. P&L',f'{month_name} consolidated P&L'),
        (f'{tab_prefix}. Orders',f'Every {month_name}-ordered transaction'),
        (f'{tab_prefix}. Refunds',f'Refunds on {month_name} orders'),
        (f'{tab_prefix}. Deferred',f'{month_name} orders still in reserve'),
        (f'{tab_prefix}. Service Fees','Ads + subscription'),
        (f'{tab_prefix}. FBA Inv Fees','Storage + prep'),
        (f'{tab_prefix}. Adjustments','Reimbursements'),
        (f'{tab_prefix}. Amazon Fees','Programme fees incl. Vine'),
        (f'{tab_prefix}. Missing','Orders requiring Seller Support'),
        (f'{tab_prefix}. Slippage','Orders with fees posted in next month'),
        (f'{tab_prefix}. SKU Rollup',f'Per-SKU {month_name} performance'),
        (None,None),('SUPPLEMENTARY',None),('Reserve Balance','Balance sheet receivable'),
    ]
    r = 5
    for item in rows_c:
        if item is None:
            r += 1; continue
        left, right = item if isinstance(item, tuple) else (item, None)
        ws.cell(row=r, column=1, value=left)
        ws.cell(row=r, column=2, value=right)
        if left and isinstance(left, str) and left.isupper() and '.' not in left:
            ws.cell(row=r, column=1).font = SECTION_FONT
        r += 1
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 65

    # P&L
    pnl_ws = wb.create_sheet(f'{tab_prefix}. P&L')
    pnl_rows = [
        (f'{month_name} {year} P&L', None, None),
        ('Every order matched to every fee paid', None, None),
        (None, None, None),
        ('VOLUME', None, None),
        ('  Total orders placed', len(target_order_ids), 'Unique order IDs'),
        ('    Matched in transactions + deferred', len(target_order_ids_seen_in_tx)+len(deferred_order_ids), None),
        ('  Order transactions', len(matched_orders), None),
        ('    Of which posted in next month', len(secondary_orders), f'{secondary_units} units, £{secondary_value:,.2f}'),
        ('  Deferred orders (in next settlement)', len(deferred_order_ids), None),
        ('  Units sold', units_by_status['Shipped'], None),
        ('  Refund transactions', len(matched_refunds), None),
        ('  Customer-cancelled orders', len(cancelled_order_ids), None),
        ('  Missing orders (Seller Support)', len(missing), None),
        (None, None, None),
        ('REVENUE', None, None),
        ('  Product sales', round(order_product_sales,2), None),
        ('  Postage credits', round(order_postage,2), None),
        ('  Gift wrap credits', round(order_giftwrap,2), None),
        ('  Product sales tax', round(order_product_tax,2), None),
        ('  Shipping credits tax', round(order_shipping_tax,2), None),
        ('  Gross revenue', round(gross_revenue,2), None),
        (None, None, None),
        ('PROMOTIONS', None, None),
        ('  Promotional rebates', round(order_promo,2), None),
        ('  Total promotions', round(total_promo,2), None),
        (None, None, None),
        ('REFUNDS', None, None),
        ('  Refund product sales', round(refund_product,2), None),
        ('  Refund postage', round(refund_postage,2), None),
        ('  Refund promo rebates', round(refund_promo,2), None),
        ('  Refund selling fees (returned)', round(refund_selling,2), None),
        ('  Refund FBA fees (returned)', round(refund_fba,2), None),
        ('  Refund other fees', round(refund_other,2), None),
        ('  Total refund net', round(total_refund_net,2), None),
        (None, None, None),
        ('NET REVENUE', round(net_revenue,2), None),
        (None, None, None),
        ('AMAZON SELLING FEES (on ordered items)', None, None),
        ('  Selling fees (referral)', round(order_selling_fees,2), None),
        ('  FBA fees', round(order_fba_fees,2), None),
        ('  Other transaction fees (DSF)', round(order_other_fees,2), None),
        ('  Total order fees', round(total_order_fees,2), None),
        (None, None, None),
        ('ADVERTISING & SUBSCRIPTION (monthly period)', None, None),
    ]
    for desc, val in sorted(service_by_desc.items()):
        pnl_rows.append((f'  {desc}', round(val,2), None))
    pnl_rows += [('  Total service fees', round(total_service,2), None), (None,None,None),
                 ('FBA INVENTORY FEES (monthly period)', None, None)]
    for desc, val in sorted(fba_inv_by_desc.items()):
        pnl_rows.append((f'  {desc}', round(val,2), None))
    pnl_rows.append(('  Total FBA inventory fees', round(total_fba_inv,2), None))
    pnl_rows.append((None,None,None))
    if adjust_by_desc:
        pnl_rows.append(('ADJUSTMENTS (monthly period)', None, None))
        for desc, val in sorted(adjust_by_desc.items()):
            pnl_rows.append((f'  {desc}', round(val,2), None))
        pnl_rows += [('  Total adjustments', round(total_adjustments,2), None), (None,None,None)]
    if amazon_fees_by_desc:
        pnl_rows.append(('AMAZON FEES (monthly period)', None, None))
        for desc, val in sorted(amazon_fees_by_desc.items()):
            pnl_rows.append((f'  {desc}', round(val,2), None))
        pnl_rows += [('  Total Amazon fees', round(total_amazon_fees,2), None), (None,None,None)]
    pnl_rows.append(('NET CONTRIBUTION', round(net_contribution,2), None))
    pnl_rows += [(None,None,None),(None,None,None),('KEY RATIOS',None,None)]
    ad_spend = abs(service_by_desc.get('Cost of Advertising', 0))
    pnl_rows += [
        ('  TACoS (ads / gross)', round(ad_spend/gross_revenue*100,2) if gross_revenue else 0, None),
        ('  Amazon take (% of gross)', round(abs(total_order_fees+total_service+total_fba_inv)/gross_revenue*100,2) if gross_revenue else 0, None),
        ('  FBA (% of gross)', round(abs(order_fba_fees)/gross_revenue*100,2) if gross_revenue else 0, None),
        ('  Selling fees (% of gross)', round(abs(order_selling_fees)/gross_revenue*100,2) if gross_revenue else 0, None),
        ('  Contribution margin (% of gross)', round(net_contribution/gross_revenue*100,2) if gross_revenue else 0, None),
        ('  Avg selling price', round(order_product_sales/units_by_status['Shipped'],2) if units_by_status['Shipped'] else 0, None),
    ]
    for i, (c1, c2, c3) in enumerate(pnl_rows, 1):
        pnl_ws.cell(row=i, column=1, value=c1)
        pnl_ws.cell(row=i, column=2, value=c2)
        pnl_ws.cell(row=i, column=3, value=c3)
        if c1 and i == 1:
            pnl_ws.cell(row=i,column=1).font = TITLE_FONT
        elif c1 and isinstance(c1,str) and c1.strip().isupper() and not c1.startswith(' '):
            pnl_ws.cell(row=i,column=1).font = SECTION_FONT
            pnl_ws.cell(row=i,column=1).fill = HEADER_FILL
        elif c1 and ('Total' in c1 or c1 == '  Gross revenue' or c1 in ('NET REVENUE','NET CONTRIBUTION')):
            pnl_ws.cell(row=i,column=1).font = TOTAL_FONT
            pnl_ws.cell(row=i,column=2).font = TOTAL_FONT
        if c1 in ('NET REVENUE','NET CONTRIBUTION'):
            pnl_ws.cell(row=i,column=1).font = Font(bold=True,size=12)
            pnl_ws.cell(row=i,column=2).font = Font(bold=True,size=12)
        if isinstance(c2, float): pnl_ws.cell(row=i,column=2).number_format = NUMBER_FMT
        elif isinstance(c2, int): pnl_ws.cell(row=i,column=2).number_format = INT_FMT
    pnl_ws.column_dimensions['A'].width = 45
    pnl_ws.column_dimensions['B'].width = 14
    pnl_ws.column_dimensions['C'].width = 35

    # Orders
    def write_detail_sheet(ws, rows, has_source=True):
        headers = ['Source','Date/Time','Settlement ID','Order ID','SKU','Qty',
                   'Product Sales','Postage','Promo','Selling Fees','FBA Fees','Other Fees','Total']
        if not has_source: headers = headers[1:]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = SECTION_FONT; cell.fill = HEADER_FILL
        for i, row in enumerate(rows, 2):
            col = 1
            if has_source:
                ws.cell(row=i, column=col, value=row.get('_source','')); col+=1
            ws.cell(row=i, column=col, value=row.get('date/time','').strip('"'));   col+=1
            ws.cell(row=i, column=col, value=row.get('settlement id','').strip('"')); col+=1
            ws.cell(row=i, column=col, value=row.get('order id','').strip('"'));    col+=1
            ws.cell(row=i, column=col, value=row.get('sku','').strip('"'));         col+=1
            ws.cell(row=i, column=col, value=to_int(row.get('quantity',0)));        col+=1
            for field in ('product sales','postage credits','promotional rebates','selling fees','fba fees','other transaction fees','total'):
                ws.cell(row=i, column=col, value=to_float(row.get(field,0))).number_format = NUMBER_FMT; col+=1
        ws.freeze_panes = 'A2'
        for col_letter, w in zip('ABCDEFGHIJKLM',[10,24,14,22,14,6,12,10,10,12,10,10,10]):
            ws.column_dimensions[col_letter].width = w

    write_detail_sheet(wb.create_sheet(f'{tab_prefix}. Orders'), matched_orders)
    write_detail_sheet(wb.create_sheet(f'{tab_prefix}. Refunds'), matched_refunds)

    # Deferred
    dws = wb.create_sheet(f'{tab_prefix}. Deferred')
    for c, h in enumerate(['Date/Time','Settlement ID','Order ID','SKU','Qty','Product Sales','Selling Fees','FBA Fees','Other Fees','Total'], 1):
        dws.cell(row=1,column=c,value=h).font = SECTION_FONT
        dws.cell(row=1,column=c).fill = HEADER_FILL
    for i, row in enumerate(matched_deferred, 2):
        dws.cell(row=i,column=1,value=row.get('date/time','').strip('"'))
        dws.cell(row=i,column=2,value=row.get('settlement id','').strip('"'))
        dws.cell(row=i,column=3,value=row.get('order id','').strip('"'))
        dws.cell(row=i,column=4,value=row.get('sku','').strip('"'))
        dws.cell(row=i,column=5,value=to_int(row.get('quantity',0)))
        for col, field in enumerate(['product sales','selling fees','fba fees','other transaction fees','total'], 6):
            dws.cell(row=i,column=col,value=to_float(row.get(field,0))).number_format = NUMBER_FMT
    dws.freeze_panes = 'A2'

    # Service Fees
    sfw = wb.create_sheet(f'{tab_prefix}. Service Fees')
    for c,h in enumerate(['Date/Time','Settlement ID','Type','Description','Other Fees','Other','Total'],1):
        sfw.cell(row=1,column=c,value=h).font=SECTION_FONT; sfw.cell(row=1,column=c).fill=HEADER_FILL
    for i,row in enumerate(period_service_fees,2):
        sfw.cell(row=i,column=1,value=row.get('date/time','').strip('"'))
        sfw.cell(row=i,column=2,value=row.get('settlement id','').strip('"'))
        sfw.cell(row=i,column=3,value=row.get('type','').strip('"'))
        sfw.cell(row=i,column=4,value=row.get('description','').strip('"'))
        sfw.cell(row=i,column=5,value=to_float(row.get('other transaction fees',0))).number_format=NUMBER_FMT
        sfw.cell(row=i,column=6,value=to_float(row.get('other',0))).number_format=NUMBER_FMT
        sfw.cell(row=i,column=7,value=to_float(row.get('total',0))).number_format=NUMBER_FMT
    sfw.freeze_panes='A2'

    # FBA Inv Fees
    fbaw = wb.create_sheet(f'{tab_prefix}. FBA Inv Fees')
    for c,h in enumerate(['Date/Time','Settlement ID','Type','Description','Total'],1):
        fbaw.cell(row=1,column=c,value=h).font=SECTION_FONT; fbaw.cell(row=1,column=c).fill=HEADER_FILL
    for i,row in enumerate(period_fba_fees,2):
        fbaw.cell(row=i,column=1,value=row.get('date/time','').strip('"'))
        fbaw.cell(row=i,column=2,value=row.get('settlement id','').strip('"'))
        fbaw.cell(row=i,column=3,value=row.get('type','').strip('"'))
        fbaw.cell(row=i,column=4,value=row.get('description','').strip('"'))
        fbaw.cell(row=i,column=5,value=to_float(row.get('total',0))).number_format=NUMBER_FMT
    fbaw.freeze_panes='A2'

    # Adjustments
    adjw = wb.create_sheet(f'{tab_prefix}. Adjustments')
    for c,h in enumerate(['Date/Time','Settlement ID','Order ID','SKU','Description','Total'],1):
        adjw.cell(row=1,column=c,value=h).font=SECTION_FONT; adjw.cell(row=1,column=c).fill=HEADER_FILL
    for i,row in enumerate(period_adjustments,2):
        adjw.cell(row=i,column=1,value=row.get('date/time','').strip('"'))
        adjw.cell(row=i,column=2,value=row.get('settlement id','').strip('"'))
        adjw.cell(row=i,column=3,value=row.get('order id','').strip('"'))
        adjw.cell(row=i,column=4,value=row.get('sku','').strip('"'))
        adjw.cell(row=i,column=5,value=row.get('description','').strip('"'))
        adjw.cell(row=i,column=6,value=to_float(row.get('total',0))).number_format=NUMBER_FMT
    adjw.freeze_panes='A2'

    # Amazon Fees
    if period_amazon_fees:
        afw = wb.create_sheet(f'{tab_prefix}. Amazon Fees')
        for c,h in enumerate(['Date/Time','Settlement ID','Order ID','SKU','Description','Total'],1):
            afw.cell(row=1,column=c,value=h).font=SECTION_FONT; afw.cell(row=1,column=c).fill=HEADER_FILL
        for i,row in enumerate(period_amazon_fees,2):
            afw.cell(row=i,column=1,value=row.get('date/time','').strip('"'))
            afw.cell(row=i,column=2,value=row.get('settlement id','').strip('"'))
            afw.cell(row=i,column=3,value=row.get('order id','').strip('"'))
            afw.cell(row=i,column=4,value=row.get('sku','').strip('"'))
            afw.cell(row=i,column=5,value=row.get('description','').strip('"'))
            afw.cell(row=i,column=6,value=to_float(row.get('total',0))).number_format=NUMBER_FMT
        afw.freeze_panes='A2'

    # Missing
    mw = wb.create_sheet(f'{tab_prefix}. Missing')
    mw.cell(row=1,column=1,value='Shipped Order ID (not in transactions or deferred)').font=SECTION_FONT
    mw.cell(row=1,column=1).fill=HEADER_FILL
    if missing:
        for i,oid in enumerate(sorted(missing),2): mw.cell(row=i,column=1,value=oid)
    else:
        mw.cell(row=2,column=1,value='— none — all shipped orders accounted for')
    mw.column_dimensions['A'].width=30

    # Slippage
    slw = wb.create_sheet(f'{tab_prefix}. Slippage')
    slw.cell(row=1,column=1,value=f'{month_name} orders with fees posted in next month').font=SECTION_FONT
    slw.cell(row=1,column=1).fill=HEADER_FILL
    slw.cell(row=2,column=1,value=f'{len(secondary_orders)} transactions, {secondary_units} units, £{secondary_value:,.2f}').font=TOTAL_FONT
    for c,h in enumerate(['Date/Time','Settlement ID','Order ID','SKU','Qty','Product Sales','Promo','Selling Fees','FBA Fees','Other','Total'],1):
        slw.cell(row=4,column=c,value=h).font=SECTION_FONT; slw.cell(row=4,column=c).fill=HEADER_FILL
    for i,row in enumerate(secondary_orders,5):
        slw.cell(row=i,column=1,value=row.get('date/time','').strip('"'))
        slw.cell(row=i,column=2,value=row.get('settlement id','').strip('"'))
        slw.cell(row=i,column=3,value=row.get('order id','').strip('"'))
        slw.cell(row=i,column=4,value=row.get('sku','').strip('"'))
        slw.cell(row=i,column=5,value=to_int(row.get('quantity',0)))
        for col,field in enumerate(['product sales','promotional rebates','selling fees','fba fees','other transaction fees','total'],6):
            slw.cell(row=i,column=col,value=to_float(row.get(field,0))).number_format=NUMBER_FMT
    slw.freeze_panes='A5'

    # SKU Rollup
    skuw = wb.create_sheet(f'{tab_prefix}. SKU Rollup')
    for c,h in enumerate(['SKU','Units','Refund Units','Product Sales','Promo','Refund Amt','Selling Fees','FBA','Other Fees','Net','Margin %','ASP'],1):
        skuw.cell(row=1,column=c,value=h).font=SECTION_FONT; skuw.cell(row=1,column=c).fill=HEADER_FILL
    for i,(sku,d) in enumerate(sorted(sku_data.items(),key=lambda kv:-kv[1]['product_sales']),2):
        net    = d['product_sales']+d['promo']+d['refund_amt']+d['selling_fees']+d['fba']+d['other']
        margin = (net/d['product_sales']*100) if d['product_sales'] else 0
        asp    = (d['product_sales']/d['units']) if d['units'] else 0
        skuw.cell(row=i,column=1,value=sku)
        skuw.cell(row=i,column=2,value=d['units']).number_format=INT_FMT
        skuw.cell(row=i,column=3,value=d['refund_units']).number_format=INT_FMT
        skuw.cell(row=i,column=4,value=round(d['product_sales'],2)).number_format=NUMBER_FMT
        skuw.cell(row=i,column=5,value=round(d['promo'],2)).number_format=NUMBER_FMT
        skuw.cell(row=i,column=6,value=round(d['refund_amt'],2)).number_format=NUMBER_FMT
        skuw.cell(row=i,column=7,value=round(d['selling_fees'],2)).number_format=NUMBER_FMT
        skuw.cell(row=i,column=8,value=round(d['fba'],2)).number_format=NUMBER_FMT
        skuw.cell(row=i,column=9,value=round(d['other'],2)).number_format=NUMBER_FMT
        skuw.cell(row=i,column=10,value=round(net,2)).number_format=NUMBER_FMT
        skuw.cell(row=i,column=11,value=round(margin,2))
        skuw.cell(row=i,column=12,value=round(asp,2)).number_format=NUMBER_FMT
    skuw.freeze_panes='A2'

    # Reserve Balance
    rbw = wb.create_sheet('Reserve Balance')
    rbw.cell(row=1,column=1,value=f'Deferred Transactions snapshot - pulled {datetime.now().strftime("%d %b %Y")}').font=SECTION_FONT
    rbw.cell(row=2,column=1,value=f'Total rows in reserve: {len(deferred)}')
    rbw.cell(row=3,column=1,value=f'Total value in reserve: £{sum(to_float(r.get("total",0)) for r in deferred):,.2f}').font=TOTAL_FONT
    for c,h in enumerate(['Date/Time','Settlement ID','Type','Order ID','SKU','Qty','Product Sales','Selling Fees','FBA Fees','Total'],1):
        rbw.cell(row=5,column=c,value=h).font=SECTION_FONT; rbw.cell(row=5,column=c).fill=HEADER_FILL
    for i,row in enumerate(deferred,6):
        rbw.cell(row=i,column=1,value=row.get('date/time','').strip('"'))
        rbw.cell(row=i,column=2,value=row.get('settlement id','').strip('"'))
        rbw.cell(row=i,column=3,value=row.get('type','').strip('"'))
        rbw.cell(row=i,column=4,value=row.get('order id','').strip('"'))
        rbw.cell(row=i,column=5,value=row.get('sku','').strip('"'))
        rbw.cell(row=i,column=6,value=to_int(row.get('quantity',0)))
        for col,field in enumerate(['product sales','selling fees','fba fees','total'],7):
            rbw.cell(row=i,column=col,value=to_float(row.get(field,0))).number_format=NUMBER_FMT
    rbw.freeze_panes='A6'

    summary = dict(
        gross_revenue=gross_revenue, total_promo=total_promo,
        total_refund_net=total_refund_net, net_revenue=net_revenue,
        total_order_fees=total_order_fees, total_service=total_service,
        total_fba_inv=total_fba_inv, total_adjustments=total_adjustments,
        total_amazon_fees=total_amazon_fees, net_contribution=net_contribution,
        shipped=len(shipped_order_ids), traced=len(target_order_ids_seen_in_tx)+len(deferred_order_ids),
        missing=len(missing), units=units_by_status['Shipped'],
    )
    return wb, summary


# ── Streamlit UI ──────────────────────────────────────────────────────────────
st.set_page_config(page_title='Amazon P&L Builder', page_icon='📦', layout='centered')
st.title('📦 Amazon P&L Builder')
st.caption('Upload your 3 Amazon reports and download a complete multi-tab P&L workbook.')

with st.form('pnl_form'):
    col1, col2 = st.columns(2)
    with col1:
        year  = st.selectbox('Year',  list(range(2023, 2028)), index=2)
    with col2:
        month = st.selectbox('Month', list(range(1, 13)),
                             format_func=lambda m: MONTH_NAMES[m], index=3)

    st.markdown('---')
    f_orders = st.file_uploader('① All Orders Report (.txt, tab-delimited)', type=['txt','csv'])
    f_tx     = st.file_uploader('② Transaction Report (.csv)',               type=['csv'])
    f_def    = st.file_uploader('③ Deferred Transactions Report (.csv)',     type=['csv'])

    submitted = st.form_submit_button('Build P&L', type='primary', use_container_width=True)

if submitted:
    if not (f_orders and f_tx and f_def):
        st.error('Please upload all three files before building.')
    else:
        with st.spinner('Processing…'):
            try:
                all_orders   = load_all_orders(f_orders)
                transactions = load_transactions(f_tx)
                deferred_    = load_deferred(f_def)
                wb, summary  = build_workbook(year, month, all_orders, transactions, deferred_)

                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)

                month_name = MONTH_NAMES[month]
                filename   = f'CSS_{month_name}_{year}_PnL.xlsx'

                st.success('✅ Workbook ready!')

                st.download_button(
                    label='⬇️  Download P&L Excel',
                    data=buf,
                    file_name=filename,
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    use_container_width=True,
                )

                st.markdown('### Headline figures')
                c1, c2, c3, c4 = st.columns(4)
                c1.metric('Gross Revenue',    f'£{summary["gross_revenue"]:,.2f}')
                c2.metric('Net Revenue',      f'£{summary["net_revenue"]:,.2f}')
                c3.metric('Net Contribution', f'£{summary["net_contribution"]:,.2f}')
                margin = summary['net_contribution']/summary['gross_revenue']*100 if summary['gross_revenue'] else 0
                c4.metric('Margin %', f'{margin:.1f}%')

                st.markdown('### Reconciliation')
                r1, r2, r3, r4 = st.columns(4)
                r1.metric('Units Sold',  summary['units'])
                r2.metric('Shipped Orders', summary['shipped'])
                r3.metric('Traced',         summary['traced'])
                r4.metric('Missing ⚠️' if summary['missing'] else 'Missing ✅', summary['missing'])

            except Exception as e:
                st.error(f'Something went wrong: {e}')
                st.exception(e)

st.markdown('---')
st.caption('Build your P&L reports')
