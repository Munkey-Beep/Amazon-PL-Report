#!/usr/bin/env python3
"""
CSS Amazon Monthly P&L Builder
Reads 3 Amazon source files (All Orders, Transactions, Deferred) and produces a
multi-tab Excel workbook matching the format established in Feb/March 2026.

Usage:
  python3 build_pnl.py \
      --month 2026-04 \
      --all-orders /path/AllOrders.txt \
      --transactions /path/Transactions.csv \
      --deferred /path/Hold.csv \
      --output /path/CSS_April_2026_PnL.xlsx
"""

import argparse
import csv
from collections import defaultdict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MONTH_NAMES = ['', 'January', 'February', 'March', 'April', 'May', 'June',
               'July', 'August', 'September', 'October', 'November', 'December']
MONTH_LETTER = {1:'Ja',2:'F',3:'M',4:'A',5:'My',6:'Jn',7:'Jl',8:'Au',9:'S',10:'O',11:'N',12:'D'}

HEADER_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True)
TOTAL_FONT = Font(bold=True)
NUMBER_FMT = '#,##0.00;[Red]-#,##0.00'
PCT_FMT = '0.00"%"'
INT_FMT = '#,##0'


def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument('--month', required=True, help='YYYY-MM, e.g. 2026-04')
    p.add_argument('--all-orders', required=True)
    p.add_argument('--transactions', required=True)
    p.add_argument('--deferred', required=True)
    p.add_argument('--output', required=True)
    return p.parse_args()


def load_all_orders(path):
    """Returns list of dicts."""
    with open(path, encoding='utf-8') as f:
        reader = csv.DictReader(f, delimiter='\t')
        return list(reader)


def load_transactions(path):
    """Skip 9 description rows, parse from row 10 (header)."""
    with open(path, encoding='utf-8-sig') as f:
        for _ in range(9):
            f.readline()
        reader = csv.DictReader(f)
        return list(reader)


def load_deferred(path):
    """Skip 7 description rows, parse from row 8 (header)."""
    with open(path, encoding='utf-8-sig') as f:
        for _ in range(7):
            f.readline()
        reader = csv.DictReader(f)
        return list(reader)


def parse_amazon_dt(s):
    """Parse '1 Apr 2026 00:06:08 UTC' -> datetime."""
    if not s:
        return None
    s = s.strip().strip('"')
    return datetime.strptime(s, '%d %b %Y %H:%M:%S UTC')


def to_float(s):
    if s is None or s == '':
        return 0.0
    if isinstance(s, (int, float)):
        return float(s)
    return float(str(s).strip().strip('"').replace(',', '') or 0)


def to_int(s):
    if s is None or s == '':
        return 0
    return int(float(str(s).strip().strip('"') or 0))


def build(args):
    year, month = map(int, args.month.split('-'))
    month_name = MONTH_NAMES[month]
    tab_prefix = MONTH_LETTER[month]

    all_orders = load_all_orders(args.all_orders)
    transactions = load_transactions(args.transactions)
    deferred = load_deferred(args.deferred)

    # === Scope: every Order ID in All Orders for the target month ===
    target_order_ids = set()
    shipped_order_ids = set()
    cancelled_order_ids = set()
    units_by_status = defaultdict(int)
    total_line_items = 0

    for row in all_orders:
        oid = row['amazon-order-id']
        target_order_ids.add(oid)
        total_line_items += 1
        status = row['order-status']
        qty = to_int(row['quantity'])
        units_by_status[status] += qty
        if status == 'Shipped':
            shipped_order_ids.add(oid)
        elif status == 'Cancelled':
            cancelled_order_ids.add(oid)

    print(f"April scope: {len(target_order_ids)} unique orders, {len(shipped_order_ids)} shipped, {len(cancelled_order_ids)} cancelled")
    print(f"Total line items: {total_line_items}, shipped units: {units_by_status['Shipped']}")

    # === Cross-reference Transactions ===
    # Each row: type can be Order, Refund, Service Fee, FBA Inventory Fee, Adjustment, Amazon Fees, Transfer
    matched_orders = []          # Order rows where order id in target scope
    matched_refunds = []         # Refund rows where order id in target scope
    period_service_fees = []     # Service Fee rows posted in target month
    period_fba_fees = []         # FBA Inventory Fee rows posted in target month
    period_adjustments = []      # Adjustment rows posted in target month
    period_amazon_fees = []      # Amazon Fees rows posted in target month
    target_order_ids_seen_in_tx = set()

    for row in transactions:
        ttype = row['type']
        oid = row.get('order id', '')
        dt = parse_amazon_dt(row.get('date/time', ''))
        posted_month = dt.month if dt else None
        posted_year = dt.year if dt else None
        is_target_month = (posted_year == year and posted_month == month)

        if ttype == 'Order' and oid in target_order_ids:
            source = 'primary' if is_target_month else 'secondary'
            row['_source'] = source
            matched_orders.append(row)
            target_order_ids_seen_in_tx.add(oid)
        elif ttype == 'Refund' and oid in target_order_ids:
            source = 'primary' if is_target_month else 'secondary'
            row['_source'] = source
            matched_refunds.append(row)
        elif ttype == 'Service Fee' and is_target_month:
            period_service_fees.append(row)
        elif ttype == 'FBA Inventory Fee' and is_target_month:
            period_fba_fees.append(row)
        elif ttype == 'Adjustment' and is_target_month:
            period_adjustments.append(row)
        elif ttype == 'Amazon Fees' and is_target_month:
            period_amazon_fees.append(row)

    # === Cross-reference Deferred ===
    matched_deferred = []
    deferred_order_ids = set()
    for row in deferred:
        oid = row.get('order id', '')
        if row.get('type') == 'Order' and oid in target_order_ids and oid not in target_order_ids_seen_in_tx:
            matched_deferred.append(row)
            deferred_order_ids.add(oid)

    # === Identify missing orders (shipped but never found anywhere) ===
    all_traced = target_order_ids_seen_in_tx | deferred_order_ids
    missing = shipped_order_ids - all_traced
    print(f"Tx matched: {len(target_order_ids_seen_in_tx)} orders ({len(matched_orders)} Order rows, {len(matched_refunds)} Refund rows)")
    print(f"Deferred matched: {len(deferred_order_ids)} orders")
    print(f"Missing (shipped but untraced): {len(missing)} orders")

    # === Aggregations ===
    def sum_col(rows, col):
        return sum(to_float(r.get(col, 0)) for r in rows)

    secondary_orders = [r for r in matched_orders if r['_source'] == 'secondary']
    secondary_units = sum(to_int(r.get('quantity', 0)) for r in secondary_orders)
    secondary_value = sum_col(secondary_orders, 'total')

    order_product_sales = sum_col(matched_orders, 'product sales')
    order_postage = sum_col(matched_orders, 'postage credits')
    order_giftwrap = sum_col(matched_orders, 'gift wrap credits')
    order_product_tax = sum_col(matched_orders, 'product sales tax')
    order_shipping_tax = sum_col(matched_orders, 'shipping credits tax')
    order_promo = sum_col(matched_orders, 'promotional rebates')
    order_selling_fees = sum_col(matched_orders, 'selling fees')
    order_fba_fees = sum_col(matched_orders, 'fba fees')
    order_other_fees = sum_col(matched_orders, 'other transaction fees')

    deferred_product = sum_col(matched_deferred, 'product sales')
    deferred_selling = sum_col(matched_deferred, 'selling fees')
    deferred_fba = sum_col(matched_deferred, 'fba fees')
    deferred_other = sum_col(matched_deferred, 'other transaction fees')

    # Refunds — sum components
    refund_product = sum_col(matched_refunds, 'product sales')
    refund_postage = sum_col(matched_refunds, 'postage credits')
    refund_promo = sum_col(matched_refunds, 'promotional rebates')
    refund_selling = sum_col(matched_refunds, 'selling fees')
    refund_fba = sum_col(matched_refunds, 'fba fees')
    refund_other = sum_col(matched_refunds, 'other transaction fees')

    gross_revenue = order_product_sales + order_postage + order_giftwrap + order_product_tax + order_shipping_tax
    total_promo = order_promo
    total_refund_net = refund_product + refund_postage + refund_promo + refund_selling + refund_fba + refund_other
    net_revenue = gross_revenue + total_promo + total_refund_net
    total_order_fees = order_selling_fees + order_fba_fees + order_other_fees

    # Period costs by description
    service_by_desc = defaultdict(float)
    for r in period_service_fees:
        desc = r.get('description', '').strip().strip('"')
        service_by_desc[desc] += to_float(r.get('total', 0))

    fba_inv_by_desc = defaultdict(float)
    for r in period_fba_fees:
        desc = r.get('description', '').strip().strip('"')
        fba_inv_by_desc[desc] += to_float(r.get('total', 0))

    adjustments_by_desc = defaultdict(float)
    for r in period_adjustments:
        desc = r.get('description', '').strip().strip('"')
        adjustments_by_desc[desc] += to_float(r.get('total', 0))

    amazon_fees_by_desc = defaultdict(float)
    for r in period_amazon_fees:
        desc = r.get('description', '').strip().strip('"')
        amazon_fees_by_desc[desc] += to_float(r.get('total', 0))

    total_service = sum(service_by_desc.values())
    total_fba_inv = sum(fba_inv_by_desc.values())
    total_adjustments = sum(adjustments_by_desc.values())
    total_amazon_fees = sum(amazon_fees_by_desc.values())

    net_contribution = net_revenue + total_order_fees + total_service + total_fba_inv + total_adjustments + total_amazon_fees

    # SKU rollup
    sku_data = defaultdict(lambda: {'units': 0, 'refund_units': 0, 'product_sales': 0.0,
                                      'promo': 0.0, 'refund_amt': 0.0, 'selling_fees': 0.0,
                                      'fba': 0.0, 'other': 0.0})
    for r in matched_orders:
        sku = r.get('sku', '').strip().strip('"') or 'NO-SKU'
        sku_data[sku]['units'] += to_int(r.get('quantity', 0))
        sku_data[sku]['product_sales'] += to_float(r.get('product sales', 0))
        sku_data[sku]['promo'] += to_float(r.get('promotional rebates', 0))
        sku_data[sku]['selling_fees'] += to_float(r.get('selling fees', 0))
        sku_data[sku]['fba'] += to_float(r.get('fba fees', 0))
        sku_data[sku]['other'] += to_float(r.get('other transaction fees', 0))
    for r in matched_refunds:
        sku = r.get('sku', '').strip().strip('"') or 'NO-SKU'
        sku_data[sku]['refund_units'] += abs(to_int(r.get('quantity', 0)))
        sku_data[sku]['refund_amt'] += to_float(r.get('product sales', 0))

    # ===== BUILD WORKBOOK =====
    wb = Workbook()
    wb.remove(wb.active)

    # --- Contents ---
    ws = wb.create_sheet('Contents')
    ws['A1'] = f'CSS Amazon UK - {month_name} {year}'
    ws['A1'].font = TITLE_FONT
    ws['A2'] = 'Complete P&L pack'
    ws['A4'] = 'CONTENTS'
    ws['A4'].font = SECTION_FONT
    contents_rows = [
        (None, None),
        ('Contents', 'This page'),
        ('Reports Used', 'What each of the three Amazon reports is and why three are needed'),
        (None, None),
        (f'{month_name.upper()} {year}', None),
        (f'{tab_prefix}. P&L', f'{month_name} consolidated P&L'),
        (f'{tab_prefix}. Orders', f'Every {month_name}-ordered transaction (primary + secondary)'),
        (f'{tab_prefix}. Refunds', f'Every refund on {month_name}-ordered items'),
        (f'{tab_prefix}. Deferred', f'{month_name} orders still in Amazon reserve'),
        (f'{tab_prefix}. Service Fees', f'{month_name} ads + subscription'),
        (f'{tab_prefix}. FBA Inv Fees', f'{month_name} storage + prep'),
        (f'{tab_prefix}. Adjustments', f'{month_name} reimbursements'),
        (f'{tab_prefix}. Amazon Fees', f'{month_name} programme fees (incl Vine)'),
        (f'{tab_prefix}. Missing', f'{month_name} orders requiring Seller Support'),
        (f'{tab_prefix}. Slippage', f'{month_name} orders with fees that posted in next month'),
        (f'{tab_prefix}. SKU Rollup', f'Per-SKU {month_name} performance'),
        (None, None),
        ('SUPPLEMENTARY', None),
        ('Reserve Balance', 'Amounts held by Amazon - balance sheet receivable'),
    ]
    r = 5
    for left, right in contents_rows:
        ws.cell(row=r, column=1, value=left)
        ws.cell(row=r, column=2, value=right)
        if left and left.isupper() and not '.' in (left or ''):
            ws.cell(row=r, column=1).font = SECTION_FONT
        r += 1
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 65

    # --- Reports Used ---
    ws = wb.create_sheet('Reports Used')
    explanations = [
        'The Amazon Reports Used to Build This P&L', None,
        'WHY MULTIPLE REPORTS ARE NEEDED',
        'No single Amazon report shows both what was ordered in a month AND all the fees paid on those orders. Orders placed in late month can ship and incur fees in the following month, so their financial picture spans two months. To get the true P&L for a calendar month of trading, three reports must be cross-referenced.',
        None,
        'REPORT 1: ALL ORDERS REPORT',
        'The definitive record of what customers ordered in a given month by purchase date. One row per order-SKU combination.',
        'WHAT IT SHOWS: Order ID, purchase date, ship date, buyer location, SKU, quantity, item price, shipping price charged to buyer, promotional discounts.',
        'WHAT IT DOES NOT SHOW: Any Amazon fees, advertising costs, FBA charges, storage, refund fees, commissions, or net cash received. It is a pre-fees top-line revenue view only.',
        'ROLE IN THIS P&L: Provides the authoritative list of Order IDs placed in the month. This defines the scope of orders to include.',
        None,
        'REPORT 2: TRANSACTION REPORT (covers target month + slippage into next month)',
        'The authoritative record of every financial transaction Amazon processed in a given date range. One row per transaction.',
        'WHAT IT SHOWS: Every order posted (with complete fee breakdown: referral fees, FBA fees, promotions, DSF, etc), every refund, every advertising charge, every storage fee, every adjustment, every bank disbursement, every debt charge. Shows Order ID and Settlement ID on each row.',
        'WHAT IT DOES NOT SHOW: Orders placed in the month that have not yet posted. A customer order placed late in the month and shipped at month-end may appear in the next month\'s report.',
        f'ROLE IN THIS P&L: Cross-referenced by Order ID against the All Orders Report. We pull a single range that covers {month_name} + the start of the following month, then split rows by posted-date into primary (in-month) and secondary (slippage).',
        None,
        'REPORT 3: DEFERRED TRANSACTIONS REPORT',
        'Shows transactions currently held in Amazon reserve awaiting release. Amazon UK holds funds for 7 days after estimated delivery before releasing them into a settlement.',
        'WHAT IT SHOWS: Every transaction still in reserve plus recently-released transactions. For each it shows which settlement ID the funds are earmarked for.',
        'ROLE IN THIS P&L: Captures any orders that shipped but are still in Amazon reserve and have not yet been posted to a transaction report. Relevant for the most recent month where some orders may still be in the 7-day hold.',
        None,
        'METHOD',
        f'For {month_name} {year}:',
        f'1. Take every Order ID from the All Orders Report ({len(target_order_ids):,} orders, {len(shipped_order_ids):,} shipped, {len(cancelled_order_ids)} cancelled).',
        '2. Cross-reference every Order ID against the Transaction Report. Flag rows posted in the target month as "primary", rows posted in the next month as "secondary" (slippage).',
        '3. Cross-reference any remaining Order IDs against the Deferred Transactions Report.',
        '4. From the Transaction Report, also pull non-order period costs posted in the target month: Service Fees, FBA Inventory Fees, Adjustments, Amazon Fees.',
        '5. Aggregate into the P&L tab.',
        f'6. Reconcile against {len(shipped_order_ids):,} shipped units.',
    ]
    r = 1
    for line in explanations:
        ws.cell(row=r, column=1, value=line)
        if line and line.isupper() and len(line) < 60:
            ws.cell(row=r, column=1).font = SECTION_FONT
        elif r == 1:
            ws.cell(row=r, column=1).font = TITLE_FONT
        r += 1
    ws.column_dimensions['A'].width = 120

    # --- P&L summary ---
    pnl_ws = wb.create_sheet(f'{tab_prefix}. P&L')
    pnl_rows = [
        (f'{month_name} {year} P&L', None, None),
        ('Every order matched to every fee paid', None, None),
        (None, None, None),
        ('VOLUME', None, None),
        ('  Total orders placed', len(target_order_ids), 'Unique order IDs in All Orders'),
        ('    Matched in transactions + deferred', len(target_order_ids_seen_in_tx) + len(deferred_order_ids), None),
        ('  Order transactions', len(matched_orders), None),
        ('    Of which posted in next month', len(secondary_orders), f'{secondary_units} units, £{secondary_value:,.2f}'),
        ('  Deferred orders (in next settlement)', len(deferred_order_ids), None),
        ('  Units sold', units_by_status['Shipped'], None),
        ('  Refund transactions', len(matched_refunds), None),
        ('  Customer-cancelled orders', len(cancelled_order_ids), None),
        ('  Missing orders (Seller Support)', len(missing), None),
        (None, None, None),
        ('REVENUE', None, None),
        ('  Product sales', round(order_product_sales, 2), None),
        ('  Postage credits', round(order_postage, 2), None),
        ('  Gift wrap credits', round(order_giftwrap, 2), None),
        ('  Product sales tax', round(order_product_tax, 2), None),
        ('  Shipping credits tax', round(order_shipping_tax, 2), None),
        ('  Gross revenue', round(gross_revenue, 2), None),
        (None, None, None),
        ('PROMOTIONS', None, None),
        ('  Promotional rebates', round(order_promo, 2), None),
        ('  Total promotions', round(total_promo, 2), None),
        (None, None, None),
        ('REFUNDS', None, None),
        ('  Refund product sales', round(refund_product, 2), None),
        ('  Refund postage', round(refund_postage, 2), None),
        ('  Refund promo rebates', round(refund_promo, 2), None),
        ('  Refund selling fees (returned)', round(refund_selling, 2), None),
        ('  Refund FBA fees (returned)', round(refund_fba, 2), None),
        ('  Refund other fees', round(refund_other, 2), None),
        ('  Total refund net', round(total_refund_net, 2), None),
        (None, None, None),
        ('NET REVENUE', round(net_revenue, 2), None),
        (None, None, None),
        ('AMAZON SELLING FEES (on ordered items)', None, None),
        ('  Selling fees (referral)', round(order_selling_fees, 2), None),
        ('  FBA fees', round(order_fba_fees, 2), None),
        ('  Other transaction fees (DSF)', round(order_other_fees, 2), None),
        ('  Total order fees', round(total_order_fees, 2), None),
        (None, None, None),
        ('ADVERTISING & SUBSCRIPTION (monthly period)', None, None),
    ]
    for desc, val in sorted(service_by_desc.items()):
        pnl_rows.append((f'  {desc}', round(val, 2), None))
    pnl_rows.append(('  Total service fees', round(total_service, 2), None))
    pnl_rows.append((None, None, None))
    pnl_rows.append(('FBA INVENTORY FEES (monthly period)', None, None))
    for desc, val in sorted(fba_inv_by_desc.items()):
        pnl_rows.append((f'  {desc}', round(val, 2), None))
    pnl_rows.append(('  Total FBA inventory fees', round(total_fba_inv, 2), None))
    pnl_rows.append((None, None, None))

    if adjustments_by_desc:
        pnl_rows.append(('ADJUSTMENTS (monthly period)', None, None))
        for desc, val in sorted(adjustments_by_desc.items()):
            pnl_rows.append((f'  {desc}', round(val, 2), None))
        pnl_rows.append(('  Total adjustments', round(total_adjustments, 2), None))
        pnl_rows.append((None, None, None))

    if amazon_fees_by_desc:
        pnl_rows.append(('AMAZON FEES (monthly period)', None, None))
        for desc, val in sorted(amazon_fees_by_desc.items()):
            pnl_rows.append((f'  {desc}', round(val, 2), None))
        pnl_rows.append(('  Total Amazon fees', round(total_amazon_fees, 2), None))
        pnl_rows.append((None, None, None))

    pnl_rows.append(('NET CONTRIBUTION', round(net_contribution, 2), None))
    pnl_rows.append((None, None, None))
    pnl_rows.append((None, None, None))
    pnl_rows.append(('KEY RATIOS', None, None))
    ad_spend = abs(service_by_desc.get('Cost of Advertising', 0))
    pnl_rows.append(('  TACoS (ads / gross)', round(ad_spend / gross_revenue * 100, 2) if gross_revenue else 0, None))
    pnl_rows.append(('  Amazon take (% of gross)', round(abs(total_order_fees + total_service + total_fba_inv) / gross_revenue * 100, 2) if gross_revenue else 0, None))
    pnl_rows.append(('  FBA (% of gross)', round(abs(order_fba_fees) / gross_revenue * 100, 2) if gross_revenue else 0, None))
    pnl_rows.append(('  Selling fees (% of gross)', round(abs(order_selling_fees) / gross_revenue * 100, 2) if gross_revenue else 0, None))
    pnl_rows.append(('  Contribution margin (% of gross)', round(net_contribution / gross_revenue * 100, 2) if gross_revenue else 0, None))
    pnl_rows.append(('  Avg selling price', round(order_product_sales / units_by_status['Shipped'], 2) if units_by_status['Shipped'] else 0, None))

    for i, (col1, col2, col3) in enumerate(pnl_rows, 1):
        pnl_ws.cell(row=i, column=1, value=col1)
        pnl_ws.cell(row=i, column=2, value=col2)
        pnl_ws.cell(row=i, column=3, value=col3)
        if col1 and i == 1:
            pnl_ws.cell(row=i, column=1).font = TITLE_FONT
        elif col1 and col1.strip().isupper() and not col1.startswith(' '):
            pnl_ws.cell(row=i, column=1).font = SECTION_FONT
            pnl_ws.cell(row=i, column=1).fill = HEADER_FILL
        elif col1 and ('Total' in col1 or col1 == '  Gross revenue' or col1 in ('NET REVENUE', 'NET CONTRIBUTION')):
            pnl_ws.cell(row=i, column=1).font = TOTAL_FONT
            pnl_ws.cell(row=i, column=2).font = TOTAL_FONT
        if col1 in ('NET REVENUE', 'NET CONTRIBUTION'):
            pnl_ws.cell(row=i, column=1).font = Font(bold=True, size=12)
            pnl_ws.cell(row=i, column=2).font = Font(bold=True, size=12)
        if isinstance(col2, float):
            pnl_ws.cell(row=i, column=2).number_format = NUMBER_FMT
        elif isinstance(col2, int):
            pnl_ws.cell(row=i, column=2).number_format = INT_FMT
    pnl_ws.column_dimensions['A'].width = 45
    pnl_ws.column_dimensions['B'].width = 14
    pnl_ws.column_dimensions['C'].width = 35

    # --- Orders detail ---
    orders_ws = wb.create_sheet(f'{tab_prefix}. Orders')
    headers = ['Source', 'Date/Time', 'Settlement ID', 'Order ID', 'SKU', 'Qty',
               'Product Sales', 'Postage', 'Promo', 'Selling Fees', 'FBA Fees', 'Other Fees', 'Total']
    for c, h in enumerate(headers, 1):
        cell = orders_ws.cell(row=1, column=c, value=h)
        cell.font = SECTION_FONT
        cell.fill = HEADER_FILL
    for i, row in enumerate(matched_orders, 2):
        orders_ws.cell(row=i, column=1, value=row['_source'])
        orders_ws.cell(row=i, column=2, value=row.get('date/time', '').strip('"'))
        orders_ws.cell(row=i, column=3, value=row.get('settlement id', '').strip('"'))
        orders_ws.cell(row=i, column=4, value=row.get('order id', '').strip('"'))
        orders_ws.cell(row=i, column=5, value=row.get('sku', '').strip('"'))
        orders_ws.cell(row=i, column=6, value=to_int(row.get('quantity', 0)))
        orders_ws.cell(row=i, column=7, value=to_float(row.get('product sales', 0))).number_format = NUMBER_FMT
        orders_ws.cell(row=i, column=8, value=to_float(row.get('postage credits', 0))).number_format = NUMBER_FMT
        orders_ws.cell(row=i, column=9, value=to_float(row.get('promotional rebates', 0))).number_format = NUMBER_FMT
        orders_ws.cell(row=i, column=10, value=to_float(row.get('selling fees', 0))).number_format = NUMBER_FMT
        orders_ws.cell(row=i, column=11, value=to_float(row.get('fba fees', 0))).number_format = NUMBER_FMT
        orders_ws.cell(row=i, column=12, value=to_float(row.get('other transaction fees', 0))).number_format = NUMBER_FMT
        orders_ws.cell(row=i, column=13, value=to_float(row.get('total', 0))).number_format = NUMBER_FMT
    for col_letter, w in zip('ABCDEFGHIJKLM', [10, 24, 14, 22, 14, 6, 12, 10, 10, 12, 10, 10, 10]):
        orders_ws.column_dimensions[col_letter].width = w
    orders_ws.freeze_panes = 'A2'

    # --- Refunds ---
    refunds_ws = wb.create_sheet(f'{tab_prefix}. Refunds')
    for c, h in enumerate(headers, 1):
        cell = refunds_ws.cell(row=1, column=c, value=h)
        cell.font = SECTION_FONT
        cell.fill = HEADER_FILL
    for i, row in enumerate(matched_refunds, 2):
        refunds_ws.cell(row=i, column=1, value=row['_source'])
        refunds_ws.cell(row=i, column=2, value=row.get('date/time', '').strip('"'))
        refunds_ws.cell(row=i, column=3, value=row.get('settlement id', '').strip('"'))
        refunds_ws.cell(row=i, column=4, value=row.get('order id', '').strip('"'))
        refunds_ws.cell(row=i, column=5, value=row.get('sku', '').strip('"'))
        refunds_ws.cell(row=i, column=6, value=to_int(row.get('quantity', 0)))
        refunds_ws.cell(row=i, column=7, value=to_float(row.get('product sales', 0))).number_format = NUMBER_FMT
        refunds_ws.cell(row=i, column=8, value=to_float(row.get('postage credits', 0))).number_format = NUMBER_FMT
        refunds_ws.cell(row=i, column=9, value=to_float(row.get('promotional rebates', 0))).number_format = NUMBER_FMT
        refunds_ws.cell(row=i, column=10, value=to_float(row.get('selling fees', 0))).number_format = NUMBER_FMT
        refunds_ws.cell(row=i, column=11, value=to_float(row.get('fba fees', 0))).number_format = NUMBER_FMT
        refunds_ws.cell(row=i, column=12, value=to_float(row.get('other transaction fees', 0))).number_format = NUMBER_FMT
        refunds_ws.cell(row=i, column=13, value=to_float(row.get('total', 0))).number_format = NUMBER_FMT
    for col_letter, w in zip('ABCDEFGHIJKLM', [10, 24, 14, 22, 14, 6, 12, 10, 10, 12, 10, 10, 10]):
        refunds_ws.column_dimensions[col_letter].width = w
    refunds_ws.freeze_panes = 'A2'

    # --- Deferred ---
    deferred_ws = wb.create_sheet(f'{tab_prefix}. Deferred')
    headers_def = ['Date/Time', 'Settlement ID', 'Order ID', 'SKU', 'Qty',
                   'Product Sales', 'Selling Fees', 'FBA Fees', 'Other Fees', 'Total']
    for c, h in enumerate(headers_def, 1):
        cell = deferred_ws.cell(row=1, column=c, value=h)
        cell.font = SECTION_FONT
        cell.fill = HEADER_FILL
    for i, row in enumerate(matched_deferred, 2):
        deferred_ws.cell(row=i, column=1, value=row.get('date/time', '').strip('"'))
        deferred_ws.cell(row=i, column=2, value=row.get('settlement id', '').strip('"'))
        deferred_ws.cell(row=i, column=3, value=row.get('order id', '').strip('"'))
        deferred_ws.cell(row=i, column=4, value=row.get('sku', '').strip('"'))
        deferred_ws.cell(row=i, column=5, value=to_int(row.get('quantity', 0)))
        deferred_ws.cell(row=i, column=6, value=to_float(row.get('product sales', 0))).number_format = NUMBER_FMT
        deferred_ws.cell(row=i, column=7, value=to_float(row.get('selling fees', 0))).number_format = NUMBER_FMT
        deferred_ws.cell(row=i, column=8, value=to_float(row.get('fba fees', 0))).number_format = NUMBER_FMT
        deferred_ws.cell(row=i, column=9, value=to_float(row.get('other transaction fees', 0))).number_format = NUMBER_FMT
        deferred_ws.cell(row=i, column=10, value=to_float(row.get('total', 0))).number_format = NUMBER_FMT
    for col_letter, w in zip('ABCDEFGHIJ', [24, 14, 22, 14, 6, 12, 12, 10, 10, 10]):
        deferred_ws.column_dimensions[col_letter].width = w
    deferred_ws.freeze_panes = 'A2'

    # --- Service Fees ---
    sf_ws = wb.create_sheet(f'{tab_prefix}. Service Fees')
    headers_sf = ['Date/Time', 'Settlement ID', 'Type', 'Description', 'Other Fees', 'Other', 'Total']
    for c, h in enumerate(headers_sf, 1):
        cell = sf_ws.cell(row=1, column=c, value=h)
        cell.font = SECTION_FONT
        cell.fill = HEADER_FILL
    for i, row in enumerate(period_service_fees, 2):
        sf_ws.cell(row=i, column=1, value=row.get('date/time', '').strip('"'))
        sf_ws.cell(row=i, column=2, value=row.get('settlement id', '').strip('"'))
        sf_ws.cell(row=i, column=3, value=row.get('type', '').strip('"'))
        sf_ws.cell(row=i, column=4, value=row.get('description', '').strip('"'))
        sf_ws.cell(row=i, column=5, value=to_float(row.get('other transaction fees', 0))).number_format = NUMBER_FMT
        sf_ws.cell(row=i, column=6, value=to_float(row.get('other', 0))).number_format = NUMBER_FMT
        sf_ws.cell(row=i, column=7, value=to_float(row.get('total', 0))).number_format = NUMBER_FMT
    for col_letter, w in zip('ABCDEFG', [24, 14, 14, 30, 12, 12, 12]):
        sf_ws.column_dimensions[col_letter].width = w
    sf_ws.freeze_panes = 'A2'

    # --- FBA Inv Fees ---
    fba_ws = wb.create_sheet(f'{tab_prefix}. FBA Inv Fees')
    headers_fba = ['Date/Time', 'Settlement ID', 'Type', 'Description', 'Total']
    for c, h in enumerate(headers_fba, 1):
        cell = fba_ws.cell(row=1, column=c, value=h)
        cell.font = SECTION_FONT
        cell.fill = HEADER_FILL
    for i, row in enumerate(period_fba_fees, 2):
        fba_ws.cell(row=i, column=1, value=row.get('date/time', '').strip('"'))
        fba_ws.cell(row=i, column=2, value=row.get('settlement id', '').strip('"'))
        fba_ws.cell(row=i, column=3, value=row.get('type', '').strip('"'))
        fba_ws.cell(row=i, column=4, value=row.get('description', '').strip('"'))
        fba_ws.cell(row=i, column=5, value=to_float(row.get('total', 0))).number_format = NUMBER_FMT
    for col_letter, w in zip('ABCDE', [24, 14, 18, 30, 12]):
        fba_ws.column_dimensions[col_letter].width = w

    # --- Adjustments ---
    adj_ws = wb.create_sheet(f'{tab_prefix}. Adjustments')
    headers_adj = ['Date/Time', 'Settlement ID', 'Order ID', 'SKU', 'Description', 'Total']
    for c, h in enumerate(headers_adj, 1):
        cell = adj_ws.cell(row=1, column=c, value=h)
        cell.font = SECTION_FONT
        cell.fill = HEADER_FILL
    for i, row in enumerate(period_adjustments, 2):
        adj_ws.cell(row=i, column=1, value=row.get('date/time', '').strip('"'))
        adj_ws.cell(row=i, column=2, value=row.get('settlement id', '').strip('"'))
        adj_ws.cell(row=i, column=3, value=row.get('order id', '').strip('"'))
        adj_ws.cell(row=i, column=4, value=row.get('sku', '').strip('"'))
        adj_ws.cell(row=i, column=5, value=row.get('description', '').strip('"'))
        adj_ws.cell(row=i, column=6, value=to_float(row.get('total', 0))).number_format = NUMBER_FMT
    for col_letter, w in zip('ABCDEF', [24, 14, 22, 14, 45, 12]):
        adj_ws.column_dimensions[col_letter].width = w

    # --- Amazon Fees ---
    if period_amazon_fees:
        af_ws = wb.create_sheet(f'{tab_prefix}. Amazon Fees')
        headers_af = ['Date/Time', 'Settlement ID', 'Order ID', 'SKU', 'Description', 'Total']
        for c, h in enumerate(headers_af, 1):
            cell = af_ws.cell(row=1, column=c, value=h)
            cell.font = SECTION_FONT
            cell.fill = HEADER_FILL
        for i, row in enumerate(period_amazon_fees, 2):
            af_ws.cell(row=i, column=1, value=row.get('date/time', '').strip('"'))
            af_ws.cell(row=i, column=2, value=row.get('settlement id', '').strip('"'))
            af_ws.cell(row=i, column=3, value=row.get('order id', '').strip('"'))
            af_ws.cell(row=i, column=4, value=row.get('sku', '').strip('"'))
            af_ws.cell(row=i, column=5, value=row.get('description', '').strip('"'))
            af_ws.cell(row=i, column=6, value=to_float(row.get('total', 0))).number_format = NUMBER_FMT
        for col_letter, w in zip('ABCDEF', [24, 14, 22, 14, 45, 12]):
            af_ws.column_dimensions[col_letter].width = w

    # --- Missing ---
    miss_ws = wb.create_sheet(f'{tab_prefix}. Missing')
    miss_ws.cell(row=1, column=1, value='Shipped Order ID (not in transactions or deferred)').font = SECTION_FONT
    miss_ws.cell(row=1, column=1).fill = HEADER_FILL
    if missing:
        for i, oid in enumerate(sorted(missing), 2):
            miss_ws.cell(row=i, column=1, value=oid)
    else:
        miss_ws.cell(row=2, column=1, value='— none — all shipped orders accounted for')
    miss_ws.column_dimensions['A'].width = 30

    # --- Slippage ---
    slip_ws = wb.create_sheet(f'{tab_prefix}. Slippage')
    slip_ws.cell(row=1, column=1, value=f'{month_name} orders with fees that posted in next month').font = SECTION_FONT
    slip_ws.cell(row=1, column=1).fill = HEADER_FILL
    slip_ws.cell(row=2, column=1, value=f'{len(secondary_orders)} transactions, {secondary_units} units, £{secondary_value:,.2f}').font = TOTAL_FONT
    slip_headers = ['Date/Time', 'Settlement ID', 'Order ID', 'SKU', 'Qty', 'Product Sales',
                    'Promo', 'Selling Fees', 'FBA Fees', 'Other', 'Total']
    for c, h in enumerate(slip_headers, 1):
        slip_ws.cell(row=4, column=c, value=h).font = SECTION_FONT
        slip_ws.cell(row=4, column=c).fill = HEADER_FILL
    for i, row in enumerate(secondary_orders, 5):
        slip_ws.cell(row=i, column=1, value=row.get('date/time', '').strip('"'))
        slip_ws.cell(row=i, column=2, value=row.get('settlement id', '').strip('"'))
        slip_ws.cell(row=i, column=3, value=row.get('order id', '').strip('"'))
        slip_ws.cell(row=i, column=4, value=row.get('sku', '').strip('"'))
        slip_ws.cell(row=i, column=5, value=to_int(row.get('quantity', 0)))
        slip_ws.cell(row=i, column=6, value=to_float(row.get('product sales', 0))).number_format = NUMBER_FMT
        slip_ws.cell(row=i, column=7, value=to_float(row.get('promotional rebates', 0))).number_format = NUMBER_FMT
        slip_ws.cell(row=i, column=8, value=to_float(row.get('selling fees', 0))).number_format = NUMBER_FMT
        slip_ws.cell(row=i, column=9, value=to_float(row.get('fba fees', 0))).number_format = NUMBER_FMT
        slip_ws.cell(row=i, column=10, value=to_float(row.get('other transaction fees', 0))).number_format = NUMBER_FMT
        slip_ws.cell(row=i, column=11, value=to_float(row.get('total', 0))).number_format = NUMBER_FMT
    for col_letter, w in zip('ABCDEFGHIJK', [24, 14, 22, 14, 6, 12, 10, 12, 10, 10, 10]):
        slip_ws.column_dimensions[col_letter].width = w
    slip_ws.freeze_panes = 'A5'

    # --- SKU Rollup ---
    sku_ws = wb.create_sheet(f'{tab_prefix}. SKU Rollup')
    sku_headers = ['SKU', 'Units', 'Refund Units', 'Product Sales', 'Promo', 'Refund Amt',
                   'Selling Fees', 'FBA', 'Other Fees', 'Net', 'Margin %', 'ASP']
    for c, h in enumerate(sku_headers, 1):
        cell = sku_ws.cell(row=1, column=c, value=h)
        cell.font = SECTION_FONT
        cell.fill = HEADER_FILL
    sku_sorted = sorted(sku_data.items(), key=lambda kv: -kv[1]['product_sales'])
    for i, (sku, d) in enumerate(sku_sorted, 2):
        net = d['product_sales'] + d['promo'] + d['refund_amt'] + d['selling_fees'] + d['fba'] + d['other']
        margin = (net / d['product_sales'] * 100) if d['product_sales'] else 0
        asp = (d['product_sales'] / d['units']) if d['units'] else 0
        sku_ws.cell(row=i, column=1, value=sku)
        sku_ws.cell(row=i, column=2, value=d['units']).number_format = INT_FMT
        sku_ws.cell(row=i, column=3, value=d['refund_units']).number_format = INT_FMT
        sku_ws.cell(row=i, column=4, value=round(d['product_sales'], 2)).number_format = NUMBER_FMT
        sku_ws.cell(row=i, column=5, value=round(d['promo'], 2)).number_format = NUMBER_FMT
        sku_ws.cell(row=i, column=6, value=round(d['refund_amt'], 2)).number_format = NUMBER_FMT
        sku_ws.cell(row=i, column=7, value=round(d['selling_fees'], 2)).number_format = NUMBER_FMT
        sku_ws.cell(row=i, column=8, value=round(d['fba'], 2)).number_format = NUMBER_FMT
        sku_ws.cell(row=i, column=9, value=round(d['other'], 2)).number_format = NUMBER_FMT
        sku_ws.cell(row=i, column=10, value=round(net, 2)).number_format = NUMBER_FMT
        sku_ws.cell(row=i, column=11, value=round(margin, 2))
        sku_ws.cell(row=i, column=12, value=round(asp, 2)).number_format = NUMBER_FMT
    for col_letter, w in zip('ABCDEFGHIJKL', [14, 8, 12, 14, 10, 12, 12, 10, 10, 12, 10, 10]):
        sku_ws.column_dimensions[col_letter].width = w
    sku_ws.freeze_panes = 'A2'

    # --- Reserve Balance (full deferred file snapshot) ---
    rb_ws = wb.create_sheet('Reserve Balance')
    rb_ws.cell(row=1, column=1, value=f'Deferred Transactions snapshot - pulled {datetime.now().strftime("%d %b %Y")}').font = SECTION_FONT
    rb_ws.cell(row=2, column=1, value=f'Total rows in reserve: {len(deferred)}')
    rb_ws.cell(row=3, column=1, value=f'Total value in reserve: £{sum(to_float(r.get("total", 0)) for r in deferred):,.2f}').font = TOTAL_FONT
    rb_headers = ['Date/Time', 'Settlement ID', 'Type', 'Order ID', 'SKU', 'Qty',
                  'Product Sales', 'Selling Fees', 'FBA Fees', 'Total']
    for c, h in enumerate(rb_headers, 1):
        cell = rb_ws.cell(row=5, column=c, value=h)
        cell.font = SECTION_FONT
        cell.fill = HEADER_FILL
    for i, row in enumerate(deferred, 6):
        rb_ws.cell(row=i, column=1, value=row.get('date/time', '').strip('"'))
        rb_ws.cell(row=i, column=2, value=row.get('settlement id', '').strip('"'))
        rb_ws.cell(row=i, column=3, value=row.get('type', '').strip('"'))
        rb_ws.cell(row=i, column=4, value=row.get('order id', '').strip('"'))
        rb_ws.cell(row=i, column=5, value=row.get('sku', '').strip('"'))
        rb_ws.cell(row=i, column=6, value=to_int(row.get('quantity', 0)))
        rb_ws.cell(row=i, column=7, value=to_float(row.get('product sales', 0))).number_format = NUMBER_FMT
        rb_ws.cell(row=i, column=8, value=to_float(row.get('selling fees', 0))).number_format = NUMBER_FMT
        rb_ws.cell(row=i, column=9, value=to_float(row.get('fba fees', 0))).number_format = NUMBER_FMT
        rb_ws.cell(row=i, column=10, value=to_float(row.get('total', 0))).number_format = NUMBER_FMT
    for col_letter, w in zip('ABCDEFGHIJ', [24, 14, 12, 22, 14, 6, 12, 12, 10, 10]):
        rb_ws.column_dimensions[col_letter].width = w
    rb_ws.freeze_panes = 'A6'

    wb.save(args.output)
    print(f"\n=== WORKBOOK SAVED ===")
    print(f"Output: {args.output}")
    print(f"\n=== HEADLINE FIGURES ===")
    print(f"Gross revenue:      £{gross_revenue:>12,.2f}")
    print(f"Promotions:         £{total_promo:>12,.2f}")
    print(f"Refunds (net):      £{total_refund_net:>12,.2f}")
    print(f"Net revenue:        £{net_revenue:>12,.2f}")
    print(f"Order fees:         £{total_order_fees:>12,.2f}")
    print(f"Service fees:       £{total_service:>12,.2f}")
    print(f"FBA inv fees:       £{total_fba_inv:>12,.2f}")
    print(f"Adjustments:        £{total_adjustments:>12,.2f}")
    print(f"Amazon fees:        £{total_amazon_fees:>12,.2f}")
    print(f"NET CONTRIBUTION:   £{net_contribution:>12,.2f}")
    print(f"\nMargin (% of gross): {net_contribution / gross_revenue * 100:.2f}%")
    print(f"\n=== RECONCILIATION ===")
    print(f"Shipped orders in All Orders:  {len(shipped_order_ids)}")
    print(f"Traced (Tx + Deferred):        {len(target_order_ids_seen_in_tx) + len(deferred_order_ids)}")
    print(f"Missing (need Seller Support): {len(missing)}")


if __name__ == '__main__':
    build(parse_args())
